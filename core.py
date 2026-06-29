#!/usr/bin/env python3
"""
iMessage Forensic Recover - Core Extraction Engine v10.0
Full-Extraction-First methodology: dump everything, parse later.
"""
TOOL_VERSION = "10.0"
import sqlite3, shutil, hashlib, re, datetime as _dt, csv, json, subprocess, sys, os
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

BANNER = """
╔══════════════════════════════════════════════════════════════════╗
║          iMessage Forensic Recover - Core Engine v9.3           ║
║          Full Extraction First · Parse Later Architecture        ║
╚══════════════════════════════════════════════════════════════════╝
"""

# ---------------------------------------------------------------------------
# Permission check
# ---------------------------------------------------------------------------

def check_tcc_permissions(db_path: Path) -> bool:
    try:
        with open(db_path, 'rb') as f:
            f.read(1)
        return True
    except PermissionError:
        print("\n[CRITICAL] Terminal lacks Full Disk Access (FDA).")
        script = (
            'display dialog "Forensic Recovery requires Full Disk Access.\\n\\n'
            'Grant access to Terminal in the Privacy Settings window, then re-run." '
            'buttons {"Open Privacy Settings", "Cancel"} default button "Open Privacy Settings" with icon stop\n'
            'if button returned of result is "Open Privacy Settings" then\n'
            '    do shell script "open \'x-apple.systempreferences:com.apple.preference.security?Privacy_AllFiles\'"\n'
            'end if'
        )
        subprocess.run(["osascript", "-e", script], check=False)
        return False
    except FileNotFoundError:
        print(f"[!] Database not found at {db_path}")
        return False

# ---------------------------------------------------------------------------
# Safe copy: db + WAL + SHM
# ---------------------------------------------------------------------------

# Prevent CSV/XLSX formula injection: a cell whose first character is one of these
# is treated as a live formula by Excel/LibreOffice/Sheets. Prefix with apostrophe.
_FORMULA_PFX = ('=', '+', '-', '@', '\t', '\r')

def _safe(v):
    s = "" if v is None else str(v)
    return ("'" + s) if (s and s[0] in _FORMULA_PFX) else v

def safe_copy_db(src_db: Path, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    copied = []
    for ext in ["", "-wal", "-shm"]:
        fpath = src_db.parent / (src_db.name + ext)
        if fpath.exists():
            if fpath.is_symlink():
                print(f"  [!] symlink detected: {fpath.name} → {fpath.resolve()} (following)")
            dest = out_dir / fpath.name
            shutil.copy2(fpath, dest)
            size_kb = dest.stat().st_size / 1024
            copied.append(f"  {fpath.name} ({size_kb:.1f} KB)")
    print("[+] Raw artifact copies:")
    for c in copied:
        print(c)
    return out_dir / src_db.name

def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()

# ---------------------------------------------------------------------------
# Manifest
# ---------------------------------------------------------------------------

def write_manifest(out_dir: Path, db_path: Path, sources: list[Path]):
    manifest = {
        "tool_version": TOOL_VERSION,
        "extraction_time": _dt.datetime.now().isoformat(),
        "source_db": str(db_path),
        "artifacts": []
    }
    for src in sources:
        if src.exists():
            manifest["artifacts"].append({
                "file": src.name,
                "size_bytes": src.stat().st_size,
                "sha256": sha256_file(src)
            })
    with open(out_dir / "MANIFEST.json", 'w') as f:
        json.dump(manifest, f, indent=2)
    print("[+] Forensic manifest written (MANIFEST.json)")

# ---------------------------------------------------------------------------
# Structured export: CSV / XLSX / JSON
# ---------------------------------------------------------------------------

MESSAGES_QUERY = """
SELECT
    m.ROWID,
    datetime(m.date/1000000000 + 978307200, 'unixepoch', 'localtime') AS date_local,
    CASE WHEN m.is_from_me = 1 THEN 'ME' ELSE h.id END              AS contact,
    m.text,
    m.is_from_me,
    m.is_delivered,
    m.is_read,
    m.is_empty,
    m.error,
    m.cache_has_attachments,
    m.message_summary_info,
    m.associated_message_guid,
    c.chat_identifier                                                 AS thread,
    a.filename                                                        AS attachment_path,
    a.mime_type                                                       AS attachment_mime
FROM message m
LEFT JOIN chat_message_join cmj  ON m.ROWID = cmj.message_id
LEFT JOIN chat c                 ON cmj.chat_id = c.ROWID
LEFT JOIN handle h               ON m.handle_id = h.ROWID
LEFT JOIN message_attachment_join maj ON m.ROWID = maj.message_id
LEFT JOIN attachment a           ON maj.attachment_id = a.ROWID
ORDER BY m.date ASC
"""

HEADERS = [
    "ROWID", "date_local", "contact", "text",
    "is_from_me", "is_delivered", "is_read", "is_empty",
    "error", "has_attachment", "summary_info", "reply_to_guid",
    "thread", "attachment_path", "attachment_mime"
]

def export_structured_data(db_path: Path, out_dir: Path) -> list[dict]:
    print("\n[*] Exporting structured message records...")
    conn = sqlite3.connect(f"file:{db_path}?mode=ro&immutable=1", uri=True)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(MESSAGES_QUERY).fetchall()
    conn.close()

    records = [dict(r) for r in rows]
    print(f"[+] {len(records)} records captured")

    # CSV
    with open(out_dir / "export.csv", 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=HEADERS)
        w.writeheader()
        for rec in records:
            w.writerow({k: _safe(v) for k, v in rec.items()})
    print("[+] export.csv written")

    # JSON (raw dump for parser.py ingestion)
    with open(out_dir / "export_raw.json", 'w', encoding='utf-8') as f:
        json.dump(records, f, indent=2, default=str)
    print("[+] export_raw.json written")

    # XLSX
    if XLSX_AVAILABLE:
        _write_xlsx(records, out_dir / "export.xlsx")
        print("[+] export.xlsx written")

    return records

def _write_xlsx(records: list[dict], path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Messages"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, rec in enumerate(records, 2):
        for col_idx, h in enumerate(HEADERS, 1):
            ws.cell(row=row_idx, column=col_idx, value=_safe(rec.get(h)))

    for i, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 4)

    wb.save(path)

# ---------------------------------------------------------------------------
# WAL carving: extract printable string fragments
# ---------------------------------------------------------------------------

WAL_HEADER_SIZE = 32
WAL_FRAME_HEADER = 24
WAL_PAGE_SIZE = 4096
TEXT_RE = re.compile(rb"[\x20-\x7E\t\n]{8,}")
# Bound resource use against a crafted WAL packed with printable bytes.
WAL_MAX_FRAGS = 50_000
WAL_MAX_FRAG_LEN = 2_000

def parse_wal(wal_path: Path, out_dir: Path):
    if not wal_path.exists():
        print("[~] No WAL file found — database may be in checkpoint state.")
        return

    print(f"\n[*] Carving WAL fragments: {wal_path.name}")
    strings: list[str] = []

    with open(wal_path, 'rb') as f:
        header = f.read(WAL_HEADER_SIZE)
        if len(header) < WAL_HEADER_SIZE:
            print("[!] WAL file too small — skipping.")
            return

        frame_num = 0
        while True:
            frame_hdr = f.read(WAL_FRAME_HEADER)
            if len(frame_hdr) < WAL_FRAME_HEADER:
                break
            page = f.read(WAL_PAGE_SIZE)
            if not page:
                break
            for match in TEXT_RE.finditer(page):
                if len(strings) >= WAL_MAX_FRAGS:
                    break
                try:
                    s = match.group(0).decode('utf-8', errors='replace').strip()
                    if 8 <= len(s) <= WAL_MAX_FRAG_LEN:
                        strings.append(s)
                except Exception:
                    pass
            frame_num += 1
            if len(strings) >= WAL_MAX_FRAGS:
                print(f"[!] WAL: fragment cap reached ({WAL_MAX_FRAGS}) — truncating carve")
                break

    print(f"[+] {frame_num} WAL frames read, {len(strings)} printable fragments extracted")

    # Deduplicated salvaged strings
    seen = set()
    unique_strings = []
    for s in strings:
        if s not in seen:
            seen.add(s)
            unique_strings.append(s)

    with open(out_dir / "salvaged_wal.txt", 'w', encoding='utf-8') as f:
        f.write(f"# WAL carve - {wal_path} - {_dt.datetime.now().isoformat()}\n")
        f.write(f"# {len(unique_strings)} unique fragments\n\n")
        for s in unique_strings:
            f.write(s + "\n")

    # Raw WAL pages as JSON for parser.py
    raw_dump = {
        "source": str(wal_path),
        "frame_count": frame_num,
        "fragment_count": len(unique_strings),
        "fragments": unique_strings
    }
    with open(out_dir / "wal_raw_dump.json", 'w', encoding='utf-8') as f:
        json.dump(raw_dump, f, indent=2)

    print("[+] salvaged_wal.txt + wal_raw_dump.json written")

# ---------------------------------------------------------------------------
# Backup scan: locate idevicebackup / iTunes backup chat.db files
# ---------------------------------------------------------------------------

# Only the real iOS backup location. We deliberately do NOT walk /var/folders:
# it is enormous on a live Mac and would open unrelated system/user SQLite caches.
BACKUP_ROOTS = [
    Path.home() / "Library" / "Application Support" / "MobileSync" / "Backup",
]

def scan_backups(out_dir: Path) -> list[Path]:
    print("\n[*] Scanning for device backup databases...")
    found: list[Path] = []
    for root in BACKUP_ROOTS:
        if not root.exists():
            continue
        for p in root.rglob("3d0d7e5fb2ce288813306e4d4636395e047a3d28"):
            found.append(p)
        # also look for chat.db-shaped databases within the backup tree
        for p in root.rglob("*.db"):
            if p.stat().st_size > 50_000:
                try:
                    c = sqlite3.connect(f"file:{p}?mode=ro&immutable=1", uri=True)
                    tables = {r[0] for r in c.execute("SELECT name FROM sqlite_master WHERE type='table'")}
                    c.close()
                    if 'message' in tables and 'handle' in tables:
                        found.append(p)
                except Exception:
                    pass

    if not found:
        print("[~] No backup databases located.")
        return []

    backup_dir = out_dir / "backups"
    backup_dir.mkdir(exist_ok=True)
    index = []
    for i, bp in enumerate(found):
        dest_name = f"backup_{i:02d}_{bp.parent.name[:8]}.db"
        shutil.copy2(bp, backup_dir / dest_name)
        index.append({"source": str(bp), "copy": dest_name})
        print(f"  [{i+1}] {bp}")

    with open(backup_dir / "backup_index.json", 'w') as f:
        json.dump(index, f, indent=2)

    print(f"[+] {len(found)} backup database(s) captured")
    return [backup_dir / e["copy"] for e in index]

# ---------------------------------------------------------------------------
# CloudKit table probe
# ---------------------------------------------------------------------------

def probe_cloudkit(db_path: Path, out_dir: Path):
    print("\n[*] Probing CloudKit sync tables...")
    try:
        conn = sqlite3.connect(f"file:{db_path}?mode=ro&immutable=1", uri=True)
        tables = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        cloudkit_tables = [t for t in tables if 'cloud' in t.lower() or 'ck' in t.lower() or 'sync' in t.lower()]

        if not cloudkit_tables:
            print("[~] No CloudKit tables detected in this database.")
            conn.close()
            return

        ck_data = {}
        for t in cloudkit_tables:
            try:
                qt = t.replace('"', '""')   # SQL-quote table name to prevent injection
                rows = conn.execute(f'SELECT * FROM "{qt}" LIMIT 500').fetchall()
                cols = [d[0] for d in conn.execute(f'SELECT * FROM "{qt}" LIMIT 1').description or []]
                ck_data[t] = {"columns": cols, "rows": [list(r) for r in rows]}
            except Exception as e:
                ck_data[t] = {"error": str(e)}

        conn.close()

        with open(out_dir / "cloudkit_probe.json", 'w') as f:
            json.dump(ck_data, f, indent=2, default=str)

        print(f"[+] {len(cloudkit_tables)} CloudKit table(s) captured: {', '.join(cloudkit_tables)}")
    except Exception as e:
        print(f"[!] CloudKit probe failed: {e}")

# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    print(BANNER)

    db_path = Path.home() / "Library" / "Messages" / "chat.db"

    if not check_tcc_permissions(db_path):
        sys.exit(1)

    ts = _dt.datetime.now().strftime('%Y%m%d_%H%M%S')
    out = Path.home() / "Desktop" / f"iMsgForensic_{ts}"
    raw_dir = out / "raw_artifacts"

    print(f"\n[*] Output folder: {out}")
    print("[*] Phase 1 — Raw Artifact Extraction\n")

    # Copy live DB + WAL + SHM
    copy = safe_copy_db(db_path, raw_dir)

    # Forensic manifest with hashes
    artifact_paths = [raw_dir / n for n in ["chat.db", "chat.db-wal", "chat.db-shm"]]
    write_manifest(raw_dir, db_path, artifact_paths)

    print("\n[*] Phase 2 — Structured Export\n")
    export_structured_data(copy, out)

    # WAL carving
    parse_wal(raw_dir / "chat.db-wal", out)

    # CloudKit probe
    probe_cloudkit(copy, out)

    # Backup scan
    scan_backups(out)

    # Summary
    total_files = sum(1 for _ in out.rglob('*') if _.is_file())
    total_bytes = sum(f.stat().st_size for f in out.rglob('*') if f.is_file())
    print(f"""
╔══════════════════════════════════════════════════════════════════╗
║  Extraction Complete
║  Output  : {out}
║  Files   : {total_files}
║  Total   : {total_bytes/1024/1024:.2f} MB
╚══════════════════════════════════════════════════════════════════╝

Next step: run parser.py against this folder for deep analysis.
  python3 parser.py --input "{out}"
""")
    input("Press Enter to exit...")

if __name__ == "__main__":
    main()
