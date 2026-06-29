#!/usr/bin/env python3
"""
iMessage Forensic Merger v1.0
Ingests multiple iMsgForensic_* output folders (one per device / backup),
deduplicates across sources, and produces a single unified chronological
timeline with per-record provenance tags.

Usage:
    # Explicit folders
    python3 merge.py --inputs ~/Desktop/iMsgForensic_20240101 ~/Desktop/iMsgForensic_20240102

    # Auto-discover all iMsgForensic_* folders on the Desktop
    python3 merge.py --scan ~/Desktop

    # Both modes work together; --output sets the destination
    python3 merge.py --scan ~/Desktop --output ~/Desktop/merged_timeline
"""
import argparse, csv, hashlib, json, re, sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

BANNER = """
╔══════════════════════════════════════════════════════════════════╗
║          iMessage Forensic Merger v1.0                          ║
║          Multi-Device · Deduplicate · Unified Timeline          ║
╚══════════════════════════════════════════════════════════════════╝
"""

# ---------------------------------------------------------------------------
# Timestamp helpers (shared with parser.py logic)
# ---------------------------------------------------------------------------

_TS_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%m/%d/%Y %H:%M",
    "%d/%m/%Y %H:%M",
    "%Y%m%d%H%M%S",
]

def parse_ts(raw) -> datetime | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        if raw > 1e15:
            raw /= 1e9
        try:
            return datetime.fromtimestamp(raw + 978307200)
        except Exception:
            return None
    s = str(raw).strip()
    for fmt in _TS_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None

def fmt_ts(dt: datetime | None) -> str:
    return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else ""

# Prevent CSV/XLSX formula injection in exported timelines.
_FORMULA_PFX = ('=', '+', '-', '@', '\t', '\r')

def _safe(v):
    s = "" if v is None else str(v)
    return ("'" + s) if (s and s[0] in _FORMULA_PFX) else v

# ---------------------------------------------------------------------------
# Source discovery
# ---------------------------------------------------------------------------

def discover_folders(scan_root: Path) -> list[Path]:
    return sorted(
        p for p in scan_root.iterdir()
        if p.is_dir() and p.name.startswith("iMsgForensic_")
    )

def label_folder(folder: Path) -> str:
    """Human-readable label: timestamp + folder name suffix."""
    return folder.name  # e.g. iMsgForensic_20240101_120000

# ---------------------------------------------------------------------------
# Load one source folder
# ---------------------------------------------------------------------------

def load_source(folder: Path) -> tuple[list[dict], list[str], dict]:
    """Returns (records, wal_fragments, manifest_info)."""
    records: list[dict] = []

    json_path = folder / "export_raw.json"
    csv_path  = folder / "export.csv"

    if json_path.exists():
        with open(json_path, encoding="utf-8") as f:
            records = json.load(f)
    elif csv_path.exists():
        with open(csv_path, encoding="utf-8") as f:
            records = list(csv.DictReader(f))
    else:
        print(f"  [!] No export found in {folder.name} — skipping records")

    wal_fragments: list[str] = []
    wal_json = folder / "wal_raw_dump.json"
    wal_txt  = folder / "salvaged_wal.txt"
    if wal_json.exists():
        with open(wal_json, encoding="utf-8") as f:
            wal_fragments = json.load(f).get("fragments", [])
    elif wal_txt.exists():
        with open(wal_txt, encoding="utf-8") as f:
            wal_fragments = [
                l.rstrip("\n") for l in f
                if not l.startswith("#") and l.strip()
            ]

    manifest: dict = {}
    mf = folder / "raw_artifacts" / "MANIFEST.json"
    if mf.exists():
        with open(mf, encoding="utf-8") as f:
            manifest = json.load(f)

    return records, wal_fragments, manifest

# ---------------------------------------------------------------------------
# Fingerprinting for deduplication
# ---------------------------------------------------------------------------

def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())

def fingerprint(record: dict) -> str:
    """
    Stable content hash across devices.
    Uses text + contact + a 60-second-bucketed timestamp so minor
    clock skew between devices doesn't create false duplicates.
    """
    ts = parse_ts(record.get("date_local") or record.get("Date"))
    if ts:
        # bucket to nearest minute
        bucket = ts.replace(second=0, microsecond=0).isoformat()
    else:
        bucket = ""

    raw = "|".join([
        _norm(record.get("text") or record.get("Text") or ""),
        _norm(record.get("contact") or record.get("Contact") or ""),
        bucket,
    ])
    return hashlib.sha1(raw.encode()).hexdigest()

# ---------------------------------------------------------------------------
# Merge engine
# ---------------------------------------------------------------------------

def merge_sources(sources: list[tuple[str, list[dict], list[str], dict]]) -> dict:
    """
    sources: list of (label, records, wal_fragments, manifest)
    Returns a result dict with unified records, WAL fragments, and stats.
    """
    # fingerprint → list of (label, record)
    seen: dict[str, list[tuple[str, dict]]] = defaultdict(list)

    for label, records, _, _ in sources:
        for rec in records:
            fp = fingerprint(rec)
            seen[fp].append((label, rec))

    unified: list[dict] = []
    conflicts: list[dict] = []

    for fp, candidates in seen.items():
        # Primary record: prefer the one from the newest source (last in list)
        primary_label, primary_rec = candidates[-1]
        ts = parse_ts(primary_rec.get("date_local") or primary_rec.get("Date"))

        merged_rec = dict(primary_rec)
        merged_rec["_fingerprint"]  = fp
        merged_rec["_ts"]           = ts
        merged_rec["_sources"]      = list({lbl for lbl, _ in candidates})
        merged_rec["_source_count"] = len(candidates)
        merged_rec["_primary_source"] = primary_label

        # Flag conflict: same fingerprint but different text across sources
        texts = {_norm(r.get("text") or "") for _, r in candidates}
        if len(texts) > 1:
            merged_rec["_conflict"] = True
            conflicts.append({
                "fingerprint": fp,
                "variants": [
                    {"source": lbl, "text": r.get("text"), "date": r.get("date_local")}
                    for lbl, r in candidates
                ]
            })
        else:
            merged_rec["_conflict"] = False

        unified.append(merged_rec)

    # Sort chronologically; null timestamps go last
    unified.sort(key=lambda r: (r["_ts"] or datetime.max))

    # Merge WAL fragments — deduplicated across all sources
    all_frags: list[str] = []
    frag_seen: set[str] = set()
    for label, _, frags, _ in sources:
        for frag in frags:
            if frag not in frag_seen:
                frag_seen.add(frag)
                all_frags.append(frag)

    # Per-source stats
    stats: dict[str, dict] = {}
    for label, records, frags, manifest in sources:
        stats[label] = {
            "records": len(records),
            "wal_fragments": len(frags),
            "manifest": manifest,
        }

    return {
        "unified": unified,
        "conflicts": conflicts,
        "wal_fragments": all_frags,
        "stats": stats,
        "total_unique": len(unified),
        "total_conflicts": len(conflicts),
    }

# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

EXPORT_HEADERS = [
    "timestamp_normalized", "contact", "text", "thread",
    "is_from_me", "is_delivered", "is_read", "is_empty",
    "error", "has_attachment", "attachment_path", "attachment_mime",
    "reply_to_guid", "ROWID",
    # provenance
    "_primary_source", "_sources", "_source_count", "_conflict", "_fingerprint",
]

def write_outputs(result: dict, out_dir: Path, source_labels: list[str]):
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    unified = result["unified"]
    conflicts = result["conflicts"]
    wal = result["wal_fragments"]

    # ── summary.json ─────────────────────────────────────────────────────────
    summary = {
        "merger_version": "1.0",
        "merged_at": datetime.now().isoformat(),
        "sources": source_labels,
        "total_unique_records": result["total_unique"],
        "total_conflicts": result["total_conflicts"],
        "total_wal_fragments": len(wal),
        "per_source": result["stats"],
    }
    with open(out_dir / "merge_summary.json", "w") as f:
        json.dump(summary, f, indent=2, default=str)

    # ── unified_timeline.csv ──────────────────────────────────────────────────
    with open(out_dir / "unified_timeline.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=EXPORT_HEADERS, extrasaction="ignore")
        w.writeheader()
        for rec in unified:
            row = dict(rec)
            row["timestamp_normalized"] = fmt_ts(rec.get("_ts"))
            row["_sources"] = "; ".join(rec.get("_sources", []))
            w.writerow({k: _safe(v) for k, v in row.items()})
    print(f"[+] unified_timeline.csv  ({len(unified)} records)")

    # ── unified_timeline.json ─────────────────────────────────────────────────
    serialisable = []
    for rec in unified:
        r = {k: v for k, v in rec.items() if k != "_ts"}
        r["timestamp_normalized"] = fmt_ts(rec.get("_ts"))
        r["_sources"] = rec.get("_sources", [])
        serialisable.append(r)
    with open(out_dir / "unified_timeline.json", "w", encoding="utf-8") as f:
        json.dump(serialisable, f, indent=2, default=str)
    print(f"[+] unified_timeline.json")

    # ── conflicts.json ────────────────────────────────────────────────────────
    with open(out_dir / "conflicts.json", "w", encoding="utf-8") as f:
        json.dump(conflicts, f, indent=2, default=str)
    if conflicts:
        print(f"[!] conflicts.json  ({len(conflicts)} records with differing text across devices)")
    else:
        print(f"[+] conflicts.json  (no conflicts)")

    # ── merged_wal_fragments.txt ──────────────────────────────────────────────
    with open(out_dir / "merged_wal_fragments.txt", "w", encoding="utf-8") as f:
        f.write(f"# Merged WAL fragments — {len(wal)} unique strings\n")
        f.write(f"# Sources: {', '.join(source_labels)}\n\n")
        for frag in wal:
            f.write(frag + "\n")
    print(f"[+] merged_wal_fragments.txt  ({len(wal)} fragments)")

    # ── XLSX ──────────────────────────────────────────────────────────────────
    if XLSX_AVAILABLE:
        _write_xlsx(unified, conflicts, out_dir / f"merged_{ts}.xlsx")

    print(f"\n[+] Output folder: {out_dir}")

def _write_xlsx(unified: list[dict], conflicts: list[dict], path: Path):
    wb = Workbook()

    def _make_sheet(name, headers, rows_fn):
        ws = wb.create_sheet(name)
        hf   = Font(color="FFFFFF", bold=True)
        hfill = PatternFill("solid", fgColor="1A4A7A")
        cf   = PatternFill("solid", fgColor="FFD0D0")  # conflict highlight
        for ci, h in enumerate(headers, 1):
            c = ws.cell(1, ci, h)
            c.font = hf
            c.fill = hfill
            c.alignment = Alignment(horizontal="center")
        for ri, row_data in enumerate(rows_fn(), 2):
            conflict = row_data.get("_conflict", False)
            for ci, h in enumerate(headers, 1):
                val = row_data.get(h)
                if isinstance(val, list):
                    val = "; ".join(str(v) for v in val)
                cell = ws.cell(ri, ci, _safe(val))
                if conflict:
                    cell.fill = cf
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)

    timeline_headers = [
        "timestamp_normalized", "contact", "text", "thread",
        "is_from_me", "has_attachment", "_primary_source", "_sources",
        "_source_count", "_conflict",
    ]

    def timeline_rows():
        for rec in unified:
            r = dict(rec)
            r["timestamp_normalized"] = fmt_ts(rec.get("_ts"))
            yield r

    _make_sheet("Unified Timeline", timeline_headers, timeline_rows)

    conflict_headers = ["fingerprint", "source", "date", "text"]
    def conflict_rows():
        for c in conflicts:
            for v in c.get("variants", []):
                yield {**v, "fingerprint": c["fingerprint"], "_conflict": True}

    _make_sheet("Conflicts", conflict_headers, conflict_rows)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(path)
    print(f"[+] {path.name}")

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    print(BANNER)

    ap = argparse.ArgumentParser(
        description="Merge multiple iMsgForensic_* output folders into a unified timeline"
    )
    ap.add_argument("--inputs", nargs="+", type=Path, default=[],
                    help="Explicit iMsgForensic_* folders to merge")
    ap.add_argument("--scan", type=Path, default=None,
                    help="Directory to scan for iMsgForensic_* folders automatically")
    ap.add_argument("--output", type=Path, default=None,
                    help="Output folder (default: <scan_dir>/merged_timeline or ./merged_timeline)")
    args = ap.parse_args()

    folders: list[Path] = list(args.inputs)

    if args.scan:
        discovered = discover_folders(args.scan)
        if not discovered:
            print(f"[!] No iMsgForensic_* folders found in {args.scan}")
        else:
            print(f"[*] Discovered {len(discovered)} folder(s) in {args.scan}:")
            for d in discovered:
                print(f"    {d.name}")
        folders += [d for d in discovered if d not in folders]

    if not folders:
        print("[!] No source folders provided. Use --inputs or --scan.")
        ap.print_help()
        sys.exit(1)

    # Validate
    valid = []
    for f in folders:
        if not f.exists():
            print(f"[!] Not found: {f} — skipping")
            continue
        valid.append(f)

    if not valid:
        print("[!] No valid source folders.")
        sys.exit(1)

    out_dir = args.output or (
        (args.scan / "merged_timeline") if args.scan else Path("merged_timeline")
    )

    print(f"\n[*] Merging {len(valid)} source(s) → {out_dir}\n")

    # Load
    sources = []
    for folder in valid:
        label = label_folder(folder)
        print(f"[*] Loading {label} ...")
        records, wal_fragments, manifest = load_source(folder)
        print(f"    {len(records)} records, {len(wal_fragments)} WAL fragments")
        sources.append((label, records, wal_fragments, manifest))

    # Merge
    print("\n[*] Deduplicating and merging...")
    result = merge_sources(sources)
    print(f"[+] {result['total_unique']} unique records across all sources")
    print(f"[+] {result['total_conflicts']} content conflicts detected")

    # Write
    print("\n[*] Writing output...\n")
    write_outputs(result, out_dir, [s[0] for s in sources])

    print(f"""
╔══════════════════════════════════════════════════════════════════╗
║  Merge complete
║  Sources : {len(valid)}
║  Records : {result['total_unique']} unique
║  Conflicts: {result['total_conflicts']}
║  WAL frags: {len(result['wal_fragments'])}
║  Output  : {out_dir}
╚══════════════════════════════════════════════════════════════════╝

Next: run parser.py against the merged folder
  python3 parser.py --input "{out_dir}"
""")

if __name__ == "__main__":
    main()
