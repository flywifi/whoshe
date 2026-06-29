#!/usr/bin/env python3
"""
iMessage Forensic Parser v1.0
Consumes output folders produced by core.py.
Normalizes timestamps, reconstructs threads, surfaces deleted tombstones,
and cross-references WAL fragments with structured records.

Usage:
    python3 parser.py --input ~/Desktop/iMsgForensic_20240101_120000
    python3 parser.py --input ~/Desktop/iMsgForensic_20240101_120000 --contact "+15551234567"
    python3 parser.py --input ~/Desktop/iMsgForensic_20240101_120000 --since 2024-01-01 --until 2024-06-01
    python3 parser.py --input ~/Desktop/iMsgForensic_20240101_120000 --keyword "meeting"
"""
import argparse, csv, json, re, sys
from pathlib import Path
from datetime import datetime, timezone
from collections import defaultdict

try:
    from extractors.attribution import attribute_fragment, build_contact_index
    _HAS_ATTRIBUTION = True
except ImportError:
    _HAS_ATTRIBUTION = False

try:
    from extractors.recursive_search import run_recursive_search
    _HAS_RECURSIVE = True
except ImportError:
    _HAS_RECURSIVE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

BANNER = """
╔══════════════════════════════════════════════════════════════════╗
║          iMessage Forensic Parser v1.0                          ║
║          Ingest · Normalize · Reconstruct · Surface             ║
╚══════════════════════════════════════════════════════════════════╝
"""

# ---------------------------------------------------------------------------
# Timestamp normalisation
# ---------------------------------------------------------------------------

_TS_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%m/%d/%Y %H:%M",
    "%d/%m/%Y %H:%M",
    "%Y%m%d%H%M%S",
]

def parse_timestamp(raw) -> datetime | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        # Apple epoch: seconds since 2001-01-01
        if raw > 1e15:
            raw /= 1e9
        if raw > 0:
            try:
                return datetime.fromtimestamp(raw + 978307200)
            except Exception:
                pass
        return None
    s = str(raw).strip()
    for fmt in _TS_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None

def fmt_ts(dt: datetime | None) -> str:
    return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else "UNKNOWN"

# Prevent CSV formula injection (=, +, -, @ trigger formula evaluation in spreadsheets).
_FORMULA_PFX = ('=', '+', '-', '@', '\t', '\r')

def _safe(v):
    s = "" if v is None else str(v)
    return ("'" + s) if (s and s[0] in _FORMULA_PFX) else v

# ---------------------------------------------------------------------------
# Data ingestion
# ---------------------------------------------------------------------------

def load_records(input_dir: Path) -> list[dict]:
    json_path = input_dir / "export_raw.json"
    csv_path = input_dir / "export.csv"

    if json_path.exists():
        print(f"[+] Loading {json_path.name}")
        with open(json_path, encoding='utf-8') as f:
            raw = json.load(f)
        records = []
        for r in raw:
            r["_ts"] = parse_timestamp(r.get("date_local") or r.get("date"))
            records.append(r)
        return records

    if csv_path.exists():
        print(f"[+] Loading {csv_path.name} (fallback)")
        records = []
        with open(csv_path, encoding='utf-8') as f:
            for row in csv.DictReader(f):
                row["_ts"] = parse_timestamp(row.get("date_local") or row.get("Date"))
                records.append(row)
        return records

    print("[!] No export_raw.json or export.csv found in input folder.")
    return []

def load_wal_fragments(input_dir: Path) -> list[str]:
    json_path = input_dir / "wal_raw_dump.json"
    txt_path = input_dir / "salvaged_wal.txt"

    if json_path.exists():
        with open(json_path, encoding='utf-8') as f:
            data = json.load(f)
        return data.get("fragments", [])

    if txt_path.exists():
        with open(txt_path, encoding='utf-8') as f:
            return [
                line.rstrip('\n')
                for line in f
                if not line.startswith('#') and line.strip()
            ]
    return []

# ---------------------------------------------------------------------------
# Filtering
# ---------------------------------------------------------------------------

def apply_filters(records: list[dict], contact: str | None, since: datetime | None,
                  until: datetime | None, keyword: str | None) -> list[dict]:
    out = []
    kw_re = re.compile(re.escape(keyword), re.IGNORECASE) if keyword else None

    for r in records:
        ts = r.get("_ts")

        if contact:
            thread = str(r.get("thread") or r.get("contact") or "")
            sender = str(r.get("contact") or "")
            if contact.lower() not in thread.lower() and contact.lower() not in sender.lower():
                continue

        if since and ts and ts < since:
            continue
        if until and ts and ts > until:
            continue

        if kw_re:
            text = str(r.get("text") or "")
            if not kw_re.search(text):
                continue

        out.append(r)
    return out

# ---------------------------------------------------------------------------
# Tombstone / deletion detection
# ---------------------------------------------------------------------------

def detect_tombstones(records: list[dict]) -> list[dict]:
    """
    Heuristic: records with is_empty=1, error != 0, or null text but
    non-null attachment are likely tombstoned (deleted) messages.
    """
    tombstones = []
    for r in records:
        reasons = []
        is_empty = str(r.get("is_empty") or "0")
        error = str(r.get("error") or "0")
        text = r.get("text")

        if is_empty == "1":
            reasons.append("is_empty=1")
        if error not in ("0", "", "None", None):
            reasons.append(f"error={error}")
        if text is None and not r.get("has_attachment") and not r.get("attachment_path"):
            reasons.append("null text, no attachment")

        if reasons:
            r["_tombstone_reasons"] = "; ".join(reasons)
            tombstones.append(r)

    return tombstones

# ---------------------------------------------------------------------------
# Thread reconstruction
# ---------------------------------------------------------------------------

def reconstruct_threads(records: list[dict]) -> dict[str, list[dict]]:
    threads: dict[str, list[dict]] = defaultdict(list)
    for r in records:
        thread_id = str(r.get("thread") or r.get("contact") or "UNKNOWN")
        threads[thread_id].append(r)
    for msgs in threads.values():
        msgs.sort(key=lambda x: (x.get("_ts") or datetime.min))
    return dict(threads)

# ---------------------------------------------------------------------------
# WAL cross-reference
# ---------------------------------------------------------------------------

_PHONE_RE = re.compile(r'(\+?1?\s*[\(\-]?\d{3}[\)\-\s]?\s*\d{3}[\-\s]?\d{4})')
_EMAIL_RE = re.compile(r'[\w\.\+\-]+@[\w\-]+\.[a-z]{2,}', re.IGNORECASE)
_MSG_BODY_RE = re.compile(r'(?:text|body|message)[=:\s]+(.{10,200})', re.IGNORECASE)

def cross_reference_wal(fragments: list[str], records: list[dict]) -> list[dict]:
    """
    Match WAL string fragments against known contacts/threads to surface
    deleted message bodies not present in the structured export.
    Also runs platform attribution when the attribution module is available.
    """
    existing_texts = {
        (str(r.get("text") or "")).strip()
        for r in records
        if r.get("text")
    }

    known_contacts = {
        (str(r.get("contact") or r.get("thread") or "")).lower()
        for r in records
    }

    contact_idx = build_contact_index(records) if _HAS_ATTRIBUTION else {}

    candidates = []
    for frag in fragments:
        if frag.strip() in existing_texts:
            continue
        if any(tok in frag for tok in ("SQLite", "rootpage", "CREATE TABLE", "PRAGMA",
                                        "bplist", "NSKeyedArchiver", "ABPerson")):
            continue
        has_contact = bool(_PHONE_RE.search(frag) or _EMAIL_RE.search(frag))
        has_body = bool(_MSG_BODY_RE.search(frag)) or (20 < len(frag) < 2000)

        if has_contact or has_body:
            ref_contacts = []
            for phone in _PHONE_RE.findall(frag):
                digits = re.sub(r'\D', '', phone)
                for c in known_contacts:
                    if digits[-7:] in re.sub(r'\D', '', c):
                        ref_contacts.append(c)

            entry: dict = {
                "fragment": frag[:500],
                "length": len(frag),
                "referenced_contacts": ref_contacts or None,
                "phones_found": _PHONE_RE.findall(frag) or None,
                "emails_found": _EMAIL_RE.findall(frag) or None,
            }

            if _HAS_ATTRIBUTION:
                attr = attribute_fragment(frag, contact_idx)
                entry["scores"]         = attr["scores"]
                entry["top_platform"]   = attr["top_platform"]
                entry["top_confidence"] = attr["top_confidence"]
                entry["is_attributed"]  = attr["is_attributed"]

            candidates.append(entry)

    return candidates

# ---------------------------------------------------------------------------
# Platform data writer
# ---------------------------------------------------------------------------

def write_platform_data(out_dir: Path, wal_candidates: list[dict]) -> dict:
    """
    Organise WAL candidates by attributed platform.
    Writes wal_attributed.json and platform_data.json.
    Returns platform_data dict.
    """
    attributed:   dict[str, list[dict]] = {}
    unattributed: list[dict] = []

    for c in wal_candidates:
        plat = c.get("top_platform")
        if plat and c.get("is_attributed"):
            attributed.setdefault(plat, []).append(c)
        else:
            unattributed.append(c)

    platform_data = {
        p: {"count": len(frags), "fragments": frags}
        for p, frags in sorted(attributed.items())
    }
    platform_data["unattributed"] = {
        "count": len(unattributed),
        "fragments": unattributed,
    }

    with open(out_dir / "wal_attributed.json", "w", encoding="utf-8") as f:
        json.dump(wal_candidates, f, indent=2, default=str)
    with open(out_dir / "platform_data.json", "w", encoding="utf-8") as f:
        json.dump(platform_data, f, indent=2, default=str)

    return platform_data


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def write_report(out_dir: Path, args, records, tombstones, threads, wal_candidates):
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')

    summary = {
        "parser_version": "1.0",
        "parsed_at": datetime.now().isoformat(),
        "source": str(args.input),
        "filters": {
            "contact": args.contact,
            "since": str(args.since) if args.since else None,
            "until": str(args.until) if args.until else None,
            "keyword": args.keyword,
        },
        "total_records": len(records),
        "tombstones_detected": len(tombstones),
        "active_threads": len(threads),
        "wal_candidates": len(wal_candidates),
    }

    with open(out_dir / "summary.json", 'w') as f:
        json.dump(summary, f, indent=2)

    # Active records
    with open(out_dir / "parsed_messages.csv", 'w', newline='', encoding='utf-8') as f:
        if records:
            fieldnames = [k for k in records[0].keys() if not k.startswith('_')]
            fieldnames.insert(0, "timestamp_normalized")
            w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            w.writeheader()
            for r in records:
                row = {k: _safe(v) for k, v in r.items() if not k.startswith('_')}
                row["timestamp_normalized"] = fmt_ts(r.get("_ts"))
                w.writerow(row)

    # Tombstones
    with open(out_dir / "tombstones.csv", 'w', newline='', encoding='utf-8') as f:
        if tombstones:
            fieldnames = ["ROWID", "timestamp_normalized", "contact", "thread",
                          "text", "error", "is_empty", "_tombstone_reasons"]
            w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            w.writeheader()
            for r in tombstones:
                row = {k: _safe(r.get(k)) for k in fieldnames}
                row["timestamp_normalized"] = fmt_ts(r.get("_ts"))
                w.writerow(row)

    # WAL candidates (base file, kept for backwards compat)
    with open(out_dir / "wal_candidates.json", 'w', encoding='utf-8') as f:
        json.dump(wal_candidates, f, indent=2, default=str)

    # Platform attribution data
    platform_data = write_platform_data(out_dir, wal_candidates)
    attributed_count = sum(
        d["count"] for k, d in platform_data.items() if k != "unattributed"
    )

    # Thread view
    with open(out_dir / "threads.json", 'w', encoding='utf-8') as f:
        serialisable = {}
        for tid, msgs in threads.items():
            serialisable[tid] = [
                {k: (fmt_ts(v) if k == "_ts" else v) for k, v in m.items()}
                for m in msgs
            ]
        json.dump(serialisable, f, indent=2, default=str)

    if XLSX_AVAILABLE:
        _write_parsed_xlsx(records, tombstones, wal_candidates, out_dir / f"parsed_{ts}.xlsx")

    # Recursive backup scan (optional)
    if _HAS_RECURSIVE:
        try:
            run_recursive_search(wal_candidates, records, args.input)
        except Exception as exc:
            print(f"[!] Recursive backup scan error: {exc}")

    plat_summary = ", ".join(
        f"{p}: {d['count']}" for p, d in platform_data.items() if p != "unattributed"
    )
    print(f"""
[+] Parser output written to: {out_dir}
    parsed_messages.csv   — {len(records)} filtered records
    tombstones.csv        — {len(tombstones)} deletion indicators
    wal_candidates.json   — {len(wal_candidates)} WAL fragments of interest
    wal_attributed.json   — {len(wal_candidates)} fragments with platform scores
    platform_data.json    — {attributed_count} attributed ({plat_summary or "none"})
    threads.json          — {len(threads)} reconstructed threads
    summary.json          — run metadata
""")

def _write_parsed_xlsx(records, tombstones, wal_candidates, path: Path):
    wb = Workbook()

    def _sheet(name, data, headers):
        ws = wb.create_sheet(name)
        hf = Font(color="FFFFFF", bold=True)
        hfill = PatternFill("solid", fgColor="2C5F8A")
        for ci, h in enumerate(headers, 1):
            c = ws.cell(1, ci, h)
            c.font = hf
            c.fill = hfill
        for ri, row in enumerate(data, 2):
            for ci, h in enumerate(headers, 1):
                ws.cell(ri, ci, _safe(row.get(h) if isinstance(row, dict) else (row[ci-1] if ci <= len(row) else None)))
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)

    msg_headers = ["timestamp_normalized", "contact", "text", "thread",
                   "is_from_me", "is_delivered", "is_read", "has_attachment"]
    msg_data = [{**{k: v for k, v in r.items() if k != "_ts"},
                 "timestamp_normalized": fmt_ts(r.get("_ts"))} for r in records]
    _sheet("Messages", msg_data, msg_headers)

    ts_headers = ["timestamp_normalized", "ROWID", "contact", "thread",
                  "text", "error", "is_empty", "_tombstone_reasons"]
    ts_data = [{**{k: v for k, v in r.items() if k != "_ts"},
                "timestamp_normalized": fmt_ts(r.get("_ts"))} for r in tombstones]
    _sheet("Tombstones", ts_data, ts_headers)

    wal_headers = ["fragment", "length", "referenced_contacts", "phones_found", "emails_found"]
    wal_data = [{h: (", ".join(v) if isinstance(v, list) and v else v)
                 for h, v in c.items()} for c in wal_candidates]
    _sheet("WAL Candidates", wal_data, wal_headers)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(path)
    print(f"[+] {path.name} written")

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_date(s: str) -> datetime:
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    raise argparse.ArgumentTypeError(f"Unrecognised date format: {s}")

def main():
    print(BANNER)
    ap = argparse.ArgumentParser(
        description="Parse iMessage forensic output folders produced by core.py"
    )
    ap.add_argument("--input",  required=True, type=Path,
                    help="Path to a Recovery_* folder produced by core.py")
    ap.add_argument("--contact", default=None,
                    help="Filter by contact phone/email/name")
    ap.add_argument("--since",  default=None, type=parse_date,
                    help="Start date YYYY-MM-DD")
    ap.add_argument("--until",  default=None, type=parse_date,
                    help="End date YYYY-MM-DD")
    ap.add_argument("--keyword", default=None,
                    help="Text keyword search (regex-safe literal)")
    ap.add_argument("--output", default=None, type=Path,
                    help="Output folder (default: <input>/parsed_output/)")
    args = ap.parse_args()

    if not args.input.exists():
        print(f"[!] Input folder not found: {args.input}")
        sys.exit(1)

    out_dir = args.output or (args.input / "parsed_output")

    print(f"[*] Input   : {args.input}")
    print(f"[*] Output  : {out_dir}")
    if args.contact: print(f"[*] Contact : {args.contact}")
    if args.since:   print(f"[*] Since   : {args.since.date()}")
    if args.until:   print(f"[*] Until   : {args.until.date()}")
    if args.keyword: print(f"[*] Keyword : {args.keyword}")

    # Load
    records = load_records(args.input)
    if not records:
        print("[!] No records loaded — check that export_raw.json or export.csv exists.")
        sys.exit(1)
    print(f"\n[*] {len(records)} total records loaded")

    wal_fragments = load_wal_fragments(args.input)
    print(f"[*] {len(wal_fragments)} WAL fragments loaded")

    # Filter
    filtered = apply_filters(records, args.contact, args.since, args.until, args.keyword)
    print(f"[*] {len(filtered)} records after filtering")

    # Analyse
    tombstones = detect_tombstones(records)
    print(f"[*] {len(tombstones)} tombstone/deletion indicators")

    threads = reconstruct_threads(filtered)
    print(f"[*] {len(threads)} conversation threads reconstructed")

    wal_candidates = cross_reference_wal(wal_fragments, records)
    print(f"[*] {len(wal_candidates)} WAL fragments flagged as potentially relevant")

    # Write
    write_report(out_dir, args, filtered, tombstones, threads, wal_candidates)

if __name__ == "__main__":
    main()
